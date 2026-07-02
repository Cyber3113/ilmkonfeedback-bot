"""
Excel hisobot (openpyxl) — 3 varaq:
1. Murojaatlar (status bo'yicha rangli qatorlar)
2. Foydalanuvchilar (murojaat soni bilan)
3. Statistika
Anonim murojaatlarda ism va telefon "—" ko'rinishida himoyalanadi.
"""
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)

STATUS_FILLS = {
    "pending": PatternFill("solid", fgColor="FFF2CC"),   # sariq
    "answered": PatternFill("solid", fgColor="D9EAD3"),  # yashil
    "closed": PatternFill("solid", fgColor="EFEFEF"),    # kulrang
}
STATUS_NAMES = {
    "pending": "Ko'rib chiqilmoqda",
    "answered": "Javob berildi",
    "closed": "Yopildi",
}
TYPE_NAMES = {"taklif": "Taklif", "etiroz": "E'tiroz"}
ROLE_NAMES = {"parent": "Ota-ona", "student": "O'quvchi", "staff": "Xodim"}


def _style_header(ws, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def export_appeals_excel(appeals, users_with_counts, stats) -> str:
    wb = Workbook()

    # ---------- 1. Murojaatlar ----------
    ws = wb.active
    ws.title = "Murojaatlar"
    headers = ["ID", "Turi", "Holati", "Ism", "Telefon", "Matn",
               "Javob", "Sana"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for a in appeals:
        name = "—" if a.is_anonymous else a.user.full_name
        phone = "—" if a.is_anonymous else a.user.phone
        row = [
            a.public_id,
            TYPE_NAMES.get(a.type, a.type),
            STATUS_NAMES.get(a.status, a.status),
            name,
            phone,
            a.text,
            a.answer or "",
            a.created_at.strftime("%d.%m.%Y %H:%M"),
        ]
        ws.append(row)
        fill = STATUS_FILLS.get(a.status)
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

    _autofit(ws, [14, 10, 18, 22, 15, 50, 40, 16])

    # ---------- 2. Foydalanuvchilar ----------
    ws2 = wb.create_sheet("Foydalanuvchilar")
    headers2 = ["Ism", "Telefon", "Roli", "Murojaatlar soni",
                "Ro'yxatdan o'tgan"]
    ws2.append(headers2)
    _style_header(ws2, len(headers2))
    for user, count in users_with_counts:
        ws2.append([
            user.full_name,
            user.phone,
            ROLE_NAMES.get(user.role, user.role),
            count,
            user.created_at.strftime("%d.%m.%Y"),
        ])
    _autofit(ws2, [25, 16, 12, 18, 18])

    # ---------- 3. Statistika ----------
    ws3 = wb.create_sheet("Statistika")
    ws3.append(["Ko'rsatkich", "Qiymat"])
    _style_header(ws3, 2)
    rows = [
        ("Jami foydalanuvchilar", stats["total_users"]),
        ("Jami murojaatlar", stats["total_appeals"]),
        ("Takliflar", stats["by_type"]["taklif"]),
        ("E'tirozlar", stats["by_type"]["etiroz"]),
        ("Ko'rib chiqilmoqda", stats["by_status"]["pending"]),
        ("Javob berildi", stats["by_status"]["answered"]),
        ("Yopildi", stats["by_status"]["closed"]),
    ]
    for r in rows:
        ws3.append(r)
    _autofit(ws3, [30, 14])

    path = f"hisobot_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(path)
    return path
